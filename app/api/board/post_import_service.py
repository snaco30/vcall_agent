from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook, load_workbook

from app.api.board.common import sanitize_html
from app.api.board_db import execute, utc_now_iso

MAX_IMPORT_ROWS = 500

TITLE_HEADERS = {"제목", "title", "Title", "TITLE"}
BODY_HEADERS = {"본문", "body", "Body", "내용", "content", "Content"}
PIN_HEADERS = {"고정", "is_pinned", "pinned", "상단고정", "pin"}
STATUS_HEADERS = {"상태", "status", "Status"}

TEMPLATE_HEADERS = ["제목", "본문", "고정", "상태"]
TEMPLATE_SAMPLE = ["공지 제목 예시", "본문 내용을 입력하세요.", "N", "published"]
TEMPLATE_FILENAME = "board_posts_import_template.xlsx"


class PostImportError(ValueError):
    """엑셀 일괄 등록 파싱·검증 오류."""


def build_import_template_xlsx() -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "게시글"
    ws.append(TEMPLATE_HEADERS)
    ws.append(TEMPLATE_SAMPLE)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 64
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), TEMPLATE_FILENAME


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value: Any) -> bool:
    text = _cell_str(value).lower()
    if not text:
        return False
    return text in {"y", "yes", "true", "1", "예", "고정", "pinned"}


def _parse_status(value: Any) -> str:
    text = _cell_str(value).lower()
    if not text or text in {"published", "publish", "발행", "공개", "y"}:
        return "published"
    if text in {"draft", "임시", "임시저장", "초안"}:
        return "draft"
    raise PostImportError(f"알 수 없는 상태값: {value}")


def _body_to_html(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    if re.search(r"<[a-z][\s\S]*>", text, re.I):
        return sanitize_html(text)
    escaped = (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )
    paragraphs = [line.strip() for line in escaped.splitlines() if line.strip()]
    if not paragraphs:
        return ""
    return sanitize_html("".join(f"<p>{part}</p>" for part in paragraphs))


def _header_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = _cell_str(header)
        if name in TITLE_HEADERS:
            mapping["title"] = index
        elif name in BODY_HEADERS:
            mapping["body"] = index
        elif name in PIN_HEADERS:
            mapping["is_pinned"] = index
        elif name in STATUS_HEADERS:
            mapping["status"] = index
    if "title" not in mapping:
        raise PostImportError("엑셀 첫 행에 '제목' 열이 필요합니다.")
    if "body" not in mapping:
        raise PostImportError("엑셀 첫 행에 '본문' 열이 필요합니다.")
    return mapping


def parse_import_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    if not file_bytes:
        raise PostImportError("업로드 파일이 비어 있습니다.")
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise PostImportError("엑셀(.xlsx) 파일만 업로드할 수 있습니다.") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise PostImportError("엑셀에 데이터가 없습니다.")

    col_map = _header_map([_cell_str(cell) for cell in rows[0]])
    parsed: list[dict[str, Any]] = []
    for offset, row in enumerate(rows[1:], start=2):
        if not row or all(_cell_str(cell) == "" for cell in row):
            continue
        title = _cell_str(row[col_map["title"]]) if col_map["title"] < len(row) else ""
        body_raw = _cell_str(row[col_map["body"]]) if col_map["body"] < len(row) else ""
        is_pinned = False
        if "is_pinned" in col_map and col_map["is_pinned"] < len(row):
            is_pinned = _parse_bool(row[col_map["is_pinned"]])
        status = "published"
        status_error = None
        if "status" in col_map and col_map["status"] < len(row):
            try:
                status = _parse_status(row[col_map["status"]])
            except PostImportError as exc:
                status_error = str(exc)
        parsed.append(
            {
                "row": offset,
                "title": title,
                "body_html": _body_to_html(body_raw),
                "is_pinned": is_pinned,
                "status": status,
                "status_error": status_error,
            }
        )
        if len(parsed) > MAX_IMPORT_ROWS:
            raise PostImportError(f"한 번에 최대 {MAX_IMPORT_ROWS}건까지 등록할 수 있습니다.")
    if not parsed:
        raise PostImportError("등록할 게시글이 없습니다. 제목·본문을 확인해 주세요.")
    return parsed


def _row_validation_errors(row: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if row.get("status_error"):
        errors.append(str(row["status_error"]))
    status = row.get("status") or "published"
    if status == "published":
        if not row.get("title"):
            errors.append("발행 글은 제목이 필요합니다.")
        if not row.get("body_html"):
            errors.append("발행 글은 본문이 필요합니다.")
    return errors


def validate_import_xlsx(file_bytes: bytes) -> dict[str, Any]:
    try:
        rows = parse_import_xlsx(file_bytes)
    except PostImportError as exc:
        return {
            "valid": False,
            "fatal_error": str(exc),
            "total_rows": 0,
            "valid_rows": 0,
            "errors": [],
            "preview": [],
            "message": str(exc),
        }

    errors: list[dict[str, Any]] = []
    valid_rows = 0
    for row in rows:
        row_errors = _row_validation_errors(row)
        if row_errors:
            for message in row_errors:
                errors.append({"row": int(row["row"]), "message": message})
        else:
            valid_rows += 1

    valid = valid_rows > 0 and not errors
    if valid:
        message = f"검증 완료: {valid_rows}건을 저장할 수 있습니다. 저장하시겠습니까?"
    elif errors:
        message = f"검증 실패: {len(errors)}개 오류가 있습니다. 파일을 수정한 뒤 다시 불러오세요."
    else:
        message = "등록할 수 있는 게시글이 없습니다."

    return {
        "valid": valid,
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "errors": errors[:30],
        "preview": [
            {
                "row": int(row["row"]),
                "title": row.get("title") or "(제목 없음)",
                "status": row.get("status") or "published",
            }
            for row in rows[:5]
        ],
        "message": message,
    }


def import_posts_from_rows(
    board_id: int,
    rows: list[dict[str, Any]],
    author_username: str,
) -> dict[str, Any]:
    imported = 0
    failed = 0
    errors: list[dict[str, Any]] = []
    post_ids: list[int] = []
    now = utc_now_iso()

    for row in rows:
        row_no = int(row["row"])
        title = row["title"]
        body_html = row["body_html"]
        status = row["status"]
        is_pinned = 1 if row["is_pinned"] else 0

        row_errors = _row_validation_errors(row)
        if row_errors:
            failed += 1
            errors.append({"row": row_no, "message": row_errors[0]})
            continue

        try:
            post_id = execute(
                """
                INSERT INTO posts(
                    board_id, title, body_html, author_username, is_pinned, view_count,
                    status, created_at, updated_at, deleted_at
                ) VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?, NULL)
                """,
                (board_id, title, body_html, author_username, is_pinned, status, now, now),
            )
            imported += 1
            post_ids.append(int(post_id))
        except Exception as exc:
            failed += 1
            errors.append({"row": row_no, "message": str(exc)[:200]})

    return {
        "imported": imported,
        "failed": failed,
        "errors": errors[:30],
        "post_ids": post_ids,
    }
